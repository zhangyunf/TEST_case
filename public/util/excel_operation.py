#!/usr/bin/env python
# -*- coding:utf-8 -*-
# Author:ZhangYunFei

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelOperation(object):

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.reader_data = {}

    def open_sheet(self, sheet_name):
        '''
        打开sheet页
        :param sheet_name: sheet页的名称
        :return:
        '''
        self.st = self.wb[sheet_name]
        return self.st

    def get_datas(self, sheet_name):
        '''
        读取数据，并
        :param sheet_name: sheet页名称
        '''
        st = self.open_sheet(sheet_name)
        for row in st.rows:
            lis = []
            for i in row:
                lis.append(i)
            self.reader_data.update({lis[0].value: lis})

    def get_readers(self):
        '''返回excel的数据'''
        return self.reader_data

    def get_value(self, cell_num):
        '''
        获取cell的值
        '''
        return cell_num.value

    def get_cell_location(self, value):
        '''返回cell的位置'''
        return "%s%d" % (value.coordinate, value.column)

    def set_cell(self, cell, color, content=None):
        '''
        标注失败和成功
        1.失败填充False且背景为红色
        2.成功填充Pass且背景为绿色
        '''
        fill = PatternFill("solid", fgColor=color)
        cell.fill = fill
        cell.value = content
        self.wb.save(self.file_path)

if __name__ == "__main__":
    fileP = r"D:\测试\自动化\TEST-case\data\case.xlsx"
    sheetN = "case"
    ex = ExcelOperation(fileP)
    ex.get_datas(sheetN)
    reader = ex.get_readers()
    for i, j in reader.items():
        for b in j:
            ex.set_cell(b, "FF0000", "测试")

